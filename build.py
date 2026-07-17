#!/usr/bin/env python3
"""
build.py — OFS Command Board weekly data builder
GE Vernova | Houston Learning Center

Usage:
  python build.py                     # auto-detect current week (Mon–Fri)
  python build.py --week 2026-06-08   # specify Monday of target week
  WEEK_OF=2026-06-08 python build.py  # same, via env (GitHub Actions)

Data sources:
  • Smartsheet API (SMARTSHEET_API_TOKEN env var required):
    Enrollment Database, Action Plan Tracker, CapEx,
    Xyleme Modernization Tracker, Exams Transfer Tracker
  • data/ xlsx files (manual upload, 3 files):
    Bowler Chart, CM Customer Demand List (OE), Class List (SS)

Generates board-data.json in repo root. Run locally or via the
"Build Command Board" GitHub Actions workflow (commits to preview).
"""

import json
import os
import re
import sys
import argparse
from datetime import date, timedelta
from pathlib import Path

# ── auto-install dependencies if needed ────────────────────────────────────
try:
    import openpyxl
except ImportError:
    import subprocess
    print("Installing openpyxl…")
    subprocess.run(
        [sys.executable, "-m", "pip", "install", "openpyxl",
         "--break-system-packages", "--quiet"],
        check=True
    )
    import openpyxl

try:
    import smartsheet
except ImportError:
    import subprocess
    print("Installing smartsheet-python-sdk…")
    subprocess.run(
        [sys.executable, "-m", "pip", "install", "smartsheet-python-sdk",
         "--break-system-packages", "--quiet"],
        check=True
    )
    import smartsheet


# ══════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# Update this section when PLLs, rules, or thresholds change.
# ══════════════════════════════════════════════════════════════════════════

# ── Smartsheet sheet IDs (pulled live at build time) ───────────────────────
SMARTSHEET_SHEETS = {
    "enrollment":    6967870172909444,   # Enrollment Database
    "actionplans":   1362792971980676,   # OFS Training Action Plan Tracker
    "capex":         3961523206573956,   # CapEx
    "modernization": 3204043720576900,   # Xyleme Training Modernization Tracker
    "exams":         8868469282918276,   # Xyleme Exams Transfer Tracker
}

# Student statuses that count as active / in-seat
ACTIVE_STATUSES = {"registered", "in progress", "completed", "auditor"}

# Student statuses to exclude
EXCLUDED_STATUSES = {
    "cancelled", "no show", "withdrawal",
    "waitlisted", "did not finish", "did not pass"
}

# Exact location strings from Enrollment Database → board category
LOCATION_MAP = {
    "houston learning center":                              "HLC",
    "birr learning center":                                 "BLC",
    "houston learning center / birr learning center":       "HLC",  # dual-site → HLC only
    "houston learning center/birr learning center":         "HLC",  # no-space variant
}

# PLL display names (key → full name)
PLL_NAMES = {
    "sherif":    "Sherif Khalifa",
    "pablo":     "Pablo Schibli",
    "ben":       "Ben Smith",
    "mohammed":  "Mohammed Nizami",
    "harry":     "Harry Hanson",
    "greg":      "Greg Walker",
    "linda":     "Linda Nelson",
}

# PLL display order (must match Slide 2 layout)
PLL_ORDER = ["sherif", "pablo", "ben", "mohammed", "harry", "greg", "linda"]

# ── Internal training routing — PROGRAM-FIRST, ORDER MATTERS ──────────────
# Each tuple: (pll_key, [keyword list])
# Keywords are checked as substrings of the lowercased class name.
# First match wins — do not reorder without updating CLAUDE.md.
INTERNAL_RULES = [
    ("harry",    ["craft", "repairs", "repair", "iles", "compressor",
                  "stator"]),
    ("sherif",   ["gas turbine", " gt ", "gt-", "ha ", "7fa", "9fa", "7ha", "9ha",
                  "eht", "frame 7", "frame 9", "osr", "combustion",
                  "mechanical crossover"]),
    ("pablo",    ["steam turbine", "hrsg", "ccpp", "wsc", "combined cycle",
                  "steam power", " st ", "blading"]),
    ("mohammed", ["controls", "control", "mkvi", "mk vi", "mkvie",
                  "fieldbus", "advant", "i&c", "gic", "simulator",
                  "site manager", "comet"]),
    ("ben",      ["excitation", "generator", "winding", "rso",
                  "surge oscillograph", "retaining ring", "end winding",
                  "biscuit", "ex2100"]),
    ("greg",     ["aeroderivative", "aero", "lm2500", "lm6000", "lm500",
                  "lms100", "aet mi"]),
    ("linda",    ["cte", "workforce", "readiness", "achieving customer"]),
]

# Vendor-led / non-technical classes — count in weekly totals only, never
# attributed to a PLL card and never flagged as unmatched (per Jim, Jul 2026).
# Matched as substrings of the lowercased class name.
VENDOR_LED_KEYWORDS = [
    "achieving customer success",
    "field engineer - onboarding",
    "train the trainer",
    "project management",
    "leadership",
]

# ── Customer (OE / SS) routing — by the TECHNOLOGY column only (per CLAUDE.md) ─
# OE = Open Enrollment (CMCustomerDemandList), SS = Site Specific (ClassList).
# Routed on the source file's Technology field; class name is a fallback only.
# Harry and Linda NEVER receive customer rows — intentionally omitted.
# Order: most-specific technologies first.
CUSTOMER_TECH_RULES = [
    ("greg",     ["aeroderivative", "lm2500", "lm6000", "lm500"]),
    ("ben",      ["excitation", "generator"]),
    ("mohammed", ["controls", "control", "gic", "simulator", "fieldbus",
                  "mark vie", "mkvie", "i&c"]),
    ("pablo",    ["steam turbine", "combined cycle", "hrsg", "ccpp", "wsc",
                  "steam power", "boiler"]),
    ("sherif",   ["gas turbine", "balance of plant", "bop", "general"]),
]

# ── Bowler KPI thresholds ──────────────────────────────────────────────────
# "plan" is the displayed target value; for reverse metrics it's also the
# green/amber boundary. Labels match the live board-data.json exactly.
BOWLER_CONFIG = {
    "ptsi": {
        "label":   "Post-Training Skill Improvement",
        "green":   70.0,
        "red":     63.0,
        "reverse": False,   # higher is better
        "plan":    70,
    },
    "timecard": {
        "label":   "Timecard On-Time Delivery",
        "green":   92.0,
        "red":     82.0,
        "reverse": False,
        "plan":    92,
    },
    "instUtil": {
        "label":   "Fully Loaded & Qualified Instructors (FLIQ)",
        "green":   15.0,    # reverse metric (lower is better) — health-monitoring only
        "red":     None,    # no red per spec
        "reverse": True,
        "plan":    15,
    },
}

# Short names for Bowler overall-RAG reason strings (e.g. "Timecard Jun 91% vs 92% plan").
BOWLER_SHORT_NAMES = {"ptsi": "PTSI", "timecard": "Timecard", "instUtil": "FLIQ"}

# ── Safety log — cumulative, always carried forward ────────────────────────
# Newest first. Update when a new advisory/incident is logged or an old one
# is retired (see CLAUDE.md "Cumulative Safety Log").
SAFETY_LOG_BASE = [
    {
        "date":      "Jul 17, 2026",
        "shortDesc": "\U0001F4F5 No-Photo Policy — Training Bays",
        "desc":      "ADVISORY — No photos in the training bays without approval. "
                     "A recent unauthorized social media post from the bays reached "
                     "senior leadership before we knew about it. No safety violation "
                     "occurred, but perception moved faster than facts. Vetted posts "
                     "are welcome — route them through HLC leadership.",
    },
    {
        "date":      "Apr 3, 2026",
        "shortDesc": "Lube oil spill — contained",
        "desc":      "Lube oil spill during maintenance exercise — contained, cleaned, no injury",
    },
    {
        "date":      "Mar 14, 2026",
        "shortDesc": "Slip — wet floor near lab",
        "desc":      "Slip on wet floor near lab entrance — no injury, area secured",
    },
]

# ── Country display-name + approximate coordinates for map pins ────────────
# Keyed by lowercased country name. Source files use many case/format variants
# (e.g. "ALGERIA", "United States Of America", "Korea, Republic of") — these are
# normalized via COUNTRY_ALIASES and canonical_country()/country_coords() so
# pins don't duplicate and don't fall back to (0,0).
COUNTRY_INFO = {
    # name (display)            (lat,    lon)
    "united states":   ("United States",        (38.0,  -97.0)),
    "mexico":          ("Mexico",               (23.6, -102.6)),
    "canada":          ("Canada",               (56.1, -106.3)),
    "brazil":          ("Brazil",               (-14.2, -51.9)),
    "argentina":       ("Argentina",            (-38.4, -63.6)),
    "chile":           ("Chile",                (-35.7, -71.5)),
    "peru":            ("Peru",                 (-9.2,  -75.0)),
    "colombia":        ("Colombia",             (4.6,   -74.1)),
    "dominican republic": ("Dominican Republic",(18.7,  -70.2)),
    "puerto rico":     ("Puerto Rico",          (18.2,  -66.5)),
    "united kingdom":  ("United Kingdom",       (55.4,  -3.4)),
    "ireland":         ("Ireland",              (53.4,  -8.2)),
    "france":          ("France",               (46.2,  2.2)),
    "germany":         ("Germany",              (51.2,  10.5)),
    "belgium":         ("Belgium",              (50.5,  4.5)),
    "netherlands":     ("Netherlands",          (52.1,  5.3)),
    "spain":           ("Spain",                (40.5,  -3.7)),
    "italy":           ("Italy",                (41.9,  12.6)),
    "switzerland":     ("Switzerland",          (46.8,  8.2)),
    "sweden":          ("Sweden",               (60.1,  18.6)),
    "norway":          ("Norway",               (60.5,  8.5)),
    "poland":          ("Poland",               (51.9,  19.1)),
    "hungary":         ("Hungary",              (47.2,  19.5)),
    "romania":         ("Romania",              (45.9,  24.9)),
    "croatia":         ("Croatia",              (45.1,  15.2)),
    "greece":          ("Greece",               (39.1,  21.8)),
    "turkey":          ("Turkey",               (38.9,  35.2)),
    "kazakhstan":      ("Kazakhstan",           (48.0,  66.9)),
    "turkmenistan":    ("Turkmenistan",         (38.9,  59.5)),
    "india":           ("India",                (20.6,  79.0)),
    "pakistan":        ("Pakistan",             (30.4,  69.3)),
    "bangladesh":      ("Bangladesh",           (23.7,  90.4)),
    "china":           ("China",                (35.9,  104.2)),
    "hong kong":       ("Hong Kong",            (22.3,  114.2)),
    "taiwan":          ("Taiwan",               (23.7,  121.0)),
    "japan":           ("Japan",                (36.2,  138.3)),
    "south korea":     ("South Korea",          (35.9,  127.8)),
    "vietnam":         ("Vietnam",              (14.1,  108.3)),
    "thailand":        ("Thailand",             (15.9,  101.0)),
    "malaysia":        ("Malaysia",             (4.2,   108.0)),
    "singapore":       ("Singapore",            (1.35,  103.8)),
    "indonesia":       ("Indonesia",            (-0.8,  113.9)),
    "philippines":     ("Philippines",          (12.9,  121.8)),
    "australia":       ("Australia",            (-25.3, 133.8)),
    "new zealand":     ("New Zealand",          (-40.9, 174.9)),
    "saudi arabia":    ("Saudi Arabia",         (24.0,  45.0)),
    "united arab emirates": ("United Arab Emirates", (23.4, 53.8)),
    "qatar":           ("Qatar",                (25.4,  51.2)),
    "bahrain":         ("Bahrain",              (26.0,  50.5)),
    "kuwait":          ("Kuwait",               (29.3,  47.5)),
    "oman":            ("Oman",                 (21.5,  55.9)),
    "iraq":            ("Iraq",                 (33.2,  43.7)),
    "jordan":          ("Jordan",               (31.2,  36.5)),
    "israel":          ("Israel",               (31.0,  34.8)),
    "egypt":           ("Egypt",                (26.8,  30.8)),
    "libya":           ("Libya",                (26.3,  17.2)),
    "tunisia":         ("Tunisia",              (33.9,  9.5)),
    "algeria":         ("Algeria",              (28.0,  3.0)),
    "morocco":         ("Morocco",              (31.8,  -7.1)),
    "senegal":         ("Senegal",              (14.5,  -14.5)),
    "cote d'ivoire":   ("Cote d'Ivoire",        (7.5,   -5.5)),
    "ghana":           ("Ghana",                (7.9,   -1.0)),
    "nigeria":         ("Nigeria",              (9.1,   8.7)),
    "angola":          ("Angola",               (-11.2, 17.9)),
    "south africa":    ("South Africa",         (-30.6, 22.9)),
}

# Variant / formal names → canonical lowercase key in COUNTRY_INFO
COUNTRY_ALIASES = {
    "united states of america":  "united states",
    "usa":                       "united states",
    "u.s.a.":                    "united states",
    "taiwan, republic of china": "taiwan",
    "taiwan republic of china":  "taiwan",
    "taiwan, province of china": "taiwan",
    "korea, republic of":        "south korea",
    "korea republic of":         "south korea",
    "republic of korea":         "south korea",
    "viet nam":                  "vietnam",
    "uae":                       "united arab emirates",
}


def _country_key(name):
    k = (name or "").strip().lower()
    return COUNTRY_ALIASES.get(k, k)


def canonical_country(name):
    """Return the clean display name for a country (dedupes case/format variants)."""
    info = COUNTRY_INFO.get(_country_key(name))
    return info[0] if info else (str(name).strip().title() if name else "Unknown")


def country_coords(name):
    """Return (lat, lon) for a country, normalizing variants; (0,0) if unknown."""
    info = COUNTRY_INFO.get(_country_key(name))
    return info[1] if info else (0.0, 0.0)


# ══════════════════════════════════════════════════════════════════════════
# UTILITIES
# ══════════════════════════════════════════════════════════════════════════

def load_existing_board(path="board-data.json"):
    """
    Load the current board-data.json (if present) so curated sections
    (CapEx, KPIs, safety, per-PLL lookAhead30) can be preserved across builds
    instead of being recomputed from raw files that don't cleanly produce them.
    Returns the parsed dict, or None if missing/unreadable.
    """
    p = Path(path)
    if not p.exists():
        return None
    try:
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return None


def find_file(patterns):
    """Find a file in data/ matching any pattern (case-insensitive substring)."""
    data_dir = Path("data")
    if not data_dir.exists():
        return None
    for f in data_dir.iterdir():
        if f.suffix.lower() != ".xlsx":
            continue
        name_lower = f.name.lower()
        for pat in patterns:
            if pat.lower() in name_lower:
                return f
    return None


def load_sheet(path, sheet_index=0):
    """Load xlsx sheet → (headers list, data rows list-of-tuples)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[sheet_index]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    return list(rows[0]), rows[1:]


# ══════════════════════════════════════════════════════════════════════════
# SMARTSHEET API ACCESS
# ══════════════════════════════════════════════════════════════════════════

def smartsheet_client():
    """Create the Smartsheet client from the SMARTSHEET_API_TOKEN env var."""
    token = os.environ.get("SMARTSHEET_API_TOKEN", "").strip()
    if not token:
        print("⛔  SMARTSHEET_API_TOKEN is not set — cannot pull live data.")
        print("    Locally: set the env var. In GitHub Actions: the workflow")
        print("    passes it from repo secrets automatically.")
        sys.exit(1)
    client = smartsheet.Smartsheet(token)
    client.errors_as_exceptions(True)
    return client


def _cell_value(cell):
    """
    Extract a plain Python value from a Smartsheet cell.
    MULTI_PICKLIST / MULTI_CONTACT cells carry their values in object_value —
    join them with ", " so downstream string matching works like the old
    xlsx exports did.
    """
    ov = getattr(cell, "object_value", None)
    if ov is not None:
        vals = getattr(ov, "values", None)
        if vals:
            try:
                return ", ".join(
                    str(getattr(v, "name", None) or getattr(v, "email", None) or v)
                    for v in vals
                )
            except TypeError:
                pass
    if cell.value is not None:
        return cell.value
    return cell.display_value


def fetch_sheet_table(ss, sheet_key, label):
    """
    Fetch a Smartsheet sheet → (headers, rows, depths).

    headers: column titles (same strings as the old xlsx export headers,
             so find_col()-based processing works unchanged)
    rows:    list of tuples of cell values in column order
    depths:  per-row hierarchy depth computed from the API's native
             parent-row links (0 = top level, 1 = child, …). This replaces
             the sheets' formula-based Hierarchy/Level columns, which go
             blank on rows without children.
    """
    print(f"  Fetching {label} from Smartsheet…")
    sheet = ss.Sheets.get_sheet(
        SMARTSHEET_SHEETS[sheet_key], level=2, include="objectValue"
    )
    headers = [c.title for c in sheet.columns]
    col_pos = {c.id: c.index for c in sheet.columns}
    n = len(headers)

    rows, depths = [], []
    depth_by_id = {}
    for row in sheet.rows:
        parent = getattr(row, "parent_id", None)
        d = depth_by_id.get(parent, -1) + 1 if parent else 0
        depth_by_id[row.id] = d
        vals = [None] * n
        for cell in row.cells:
            i = col_pos.get(cell.column_id)
            if i is not None:
                vals[i] = _cell_value(cell)
        rows.append(tuple(vals))
        depths.append(d)
    print(f"    → {len(rows)} rows")
    return headers, rows, depths


def find_col(headers, *candidates):
    """
    Find 0-based column index by header name. EXACT matches win over substring
    matches (so "location" finds the "Location" column, not "Delivery Location"),
    and earlier candidates win over later ones.
    """
    hl = [str(h).lower().strip() if h is not None else "" for h in headers]
    # Pass 1 — exact match (most specific)
    for candidate in candidates:
        cl = candidate.lower().strip()
        for i, h in enumerate(hl):
            if h == cl:
                return i
    # Pass 2 — substring fallback
    for candidate in candidates:
        cl = candidate.lower().strip()
        for i, h in enumerate(hl):
            if cl in h:
                return i
    return None


def to_date(val):
    """Convert openpyxl date value to Python date."""
    if val is None:
        return None
    if hasattr(val, "date"):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        from datetime import datetime
        s = val.strip()
        if "T" in s:               # ISO datetime from the Smartsheet API
            s = s.split("T")[0]
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def classify_location(raw):
    """Map raw location string to HLC / BLC / KLC / Other."""
    if not raw:
        return "Other"
    s = str(raw).strip()
    key = s.lower()
    if key in LOCATION_MAP:
        return LOCATION_MAP[key]
    # Fallback substring checks
    if "houston" in key:
        return "HLC"
    if "birr" in key:
        return "BLC"
    if "kuwait" in key:
        return "KLC"
    return "Other"


def route_pll(class_name):
    """
    Route an INTERNAL class (from the Enrollment Database) to a PLL using
    program/technology keyword matching on the class name. Craft always →
    Harry. First match wins. Returns (pll_key or None, flagged bool).

    NOTE: every Enrollment Database row is INTERNAL training regardless of its
    delivery location (HLC/BLC/KLC/Other) — bucket is determined by source file,
    not location — so this always uses INTERNAL_RULES.
    """
    name = " " + class_name.lower() + " "

    # Vendor-led / non-technical classes: totals only, no PLL attribution.
    for kw in VENDOR_LED_KEYWORDS:
        if kw in name:
            return "__vendor__", False

    # CTE check MUST precede the craft rule — CTE-program classes
    # (e.g. "Craft Entry Level CTE Program") belong to Linda, not Harry.
    for kw in (" cte ", "cte program", "workforce", "readiness"):
        if kw in name:
            return "linda", False

    # Craft classes otherwise always route to Harry (he owns all craft training).
    if "craft" in name:
        return "harry", False

    for pll_key, keywords in INTERNAL_RULES:
        for kw in keywords:
            if kw in name:
                return pll_key, False

    return None, True  # unmatched — flag to Jim


def route_customer(technology, class_name=""):
    """
    Route an OE / SS (customer) class to a PLL by the TECHNOLOGY column only
    (per CLAUDE.md). Falls back to class-name tokens only when Technology is
    blank/unrecognized. Harry and Linda never receive customer rows.
    Returns (pll_key or None, flagged bool).
    """
    tech = (technology or "").strip().lower()
    if tech:
        for pll_key, keywords in CUSTOMER_TECH_RULES:
            for kw in keywords:
                if kw in tech:
                    return pll_key, False
    name = (class_name or "").lower()
    if name:
        for pll_key, keywords in CUSTOMER_TECH_RULES:
            for kw in keywords:
                if kw in name:
                    return pll_key, False
    return None, True  # unmatched — flag to Jim


def bowler_rag(key, value):
    """
    Calculate RAG for a single Bowler KPI value.
    Reverse metrics (lower is better, e.g. FLIQ) have no red — green below
    the threshold, amber at or above it.
    """
    cfg = BOWLER_CONFIG.get(key, {})
    if value is None or cfg.get("green") is None:
        return "amber"
    if cfg.get("reverse"):
        return "green" if value < cfg["green"] else "amber"
    if value >= cfg["green"]:
        return "green"
    if cfg.get("red") is not None and value < cfg["red"]:
        return "red"
    return "amber"


def safety_kpi_rag(value):
    """Binary RAG for a safety KPI vs its 90% plan — green >= 90, else red."""
    if value is None:
        return "amber"
    return "green" if value >= 90 else "red"


def worst_rag(*rags):
    """Return the worst RAG from a list."""
    priority = {"red": 0, "amber": 1, "green": 2}
    return min(rags, key=lambda r: priority.get(r, 1))


# ══════════════════════════════════════════════════════════════════════════
# WEEK DETECTION
# ══════════════════════════════════════════════════════════════════════════

def sanitize_week_override(raw):
    """
    Clean a raw --week / WEEK_OF value before parsing as an ISO date: strip
    surrounding whitespace/quotes and a leading "week_of:" or "week_of="
    label (case-insensitive). A pasted 'week_of: 2026-07-20' crashed run #1
    on 2026-07-17.
    """
    s = raw.strip().strip("'\"").strip()
    s = re.sub(r"(?i)^week_of\s*[:=]\s*", "", s)
    return s.strip().strip("'\"").strip()


def get_week(override=None):
    """Return (week_start, week_end) as date objects (Mon–Fri)."""
    if override:
        cleaned = sanitize_week_override(override)
        try:
            ws = date.fromisoformat(cleaned)
        except ValueError:
            print(f"⛔  --week/WEEK_OF value {override!r} is not a valid date "
                  f"after cleanup (got {cleaned!r}). Expected format: "
                  f"YYYY-MM-DD — the Monday of the target week, "
                  f"e.g. 2026-07-20.")
            sys.exit(1)
    else:
        today = date.today()
        ws = today - timedelta(days=today.weekday())  # Monday
    return ws, ws + timedelta(days=4)


def format_week_label(ws, we):
    """Format as 'Jun 8 – 12, 2026'."""
    if ws.month == we.month:
        return f"{ws.strftime('%b')} {ws.day} \u2013 {we.day}, {we.year}"
    return f"{ws.strftime('%b')} {ws.day} \u2013 {we.strftime('%b')} {we.day}, {we.year}"


# ══════════════════════════════════════════════════════════════════════════
# FILE DISCOVERY
# ══════════════════════════════════════════════════════════════════════════

# Manual xlsx uploads — everything else comes live from the Smartsheet API.
SOURCE_FILES = {
    "demand":    ["cmcustomerdemandlist", "cm customer"],
    "classlist": ["classlist"],
    "bowler":    ["bowler chart"],
}


def discover_files():
    """Find the 3 manual source files. Returns (found dict, missing list)."""
    found, missing = {}, []
    for key, patterns in SOURCE_FILES.items():
        f = find_file(patterns)
        if f:
            found[key] = f
        else:
            missing.append(key)
    return found, missing


# ══════════════════════════════════════════════════════════════════════════
# ENROLLMENT PROCESSING
# ══════════════════════════════════════════════════════════════════════════

def process_enrollment(headers, rows, week_start, week_end):
    """
    Process the Enrollment Database (live Smartsheet pull) and return all
    data needed for the board. Counting rules are unchanged from the xlsx
    era: in-session window (Start <= week_end AND Finish >= week_start) and
    per-student active-status filtering — see CLAUDE.md.

    Returns dict with:
      location_counts, pll_classes, flags,
      total_internal, total_oe, internal_courses, oe_courses,
      country_counts
    """
    print("  Processing Enrollment Database…")

    # Detect columns by header name
    c = {
        "course":   find_col(headers, "course name", "class name", "course"),
        "start":    find_col(headers, "start date", "class start", "start"),
        "finish":   find_col(headers, "finish date", "end date", "class end", "finish"),
        "location": find_col(headers, "location", "site"),
        "status":   find_col(headers, "student status", "enrollment status", "status"),
        "country":  find_col(headers, "student country", "country"),
        "capacity": find_col(headers, "capacity", "class capacity", "max capacity"),
    }

    missing_cols = [k for k, v in c.items() if v is None
                    and k not in ("capacity",)]
    if missing_cols:
        print(f"  ⚠️  Column(s) not found: {missing_cols}")
        print(f"      Available: {[h for h in headers if h][:20]}")

    def get(row, col_key):
        idx = c.get(col_key)
        return row[idx] if idx is not None and idx < len(row) else None

    # Aggregate rows by class name
    # class_key = (course_name, location) to handle same-name classes at diff sites
    classes = {}

    for row in rows:
        course = get(row, "course")
        if not course:
            continue
        course = str(course).strip()

        start  = to_date(get(row, "start"))
        finish = to_date(get(row, "finish"))
        if not start or not finish:
            continue

        # In-session filter: class overlaps with this week
        if start > week_end or finish < week_start:
            continue

        status = str(get(row, "status") or "").strip().lower()
        if status not in ACTIVE_STATUSES:
            continue

        location = str(get(row, "location") or "").strip()
        country  = canonical_country(get(row, "country"))
        capacity = get(row, "capacity")
        loc_cat  = classify_location(location)

        key = (course, loc_cat)
        if key not in classes:
            classes[key] = {
                "name":     course,
                "loc_cat":  loc_cat,
                "capacity": int(capacity) if capacity else 0,
                "students": 0,
                "countries": {},
            }

        classes[key]["students"] += 1
        if country:
            classes[key]["countries"][country] = \
                classes[key]["countries"].get(country, 0) + 1

    # Route each class to PLL
    location_counts  = {"HLC": 0, "BLC": 0, "KLC": 0, "Other": 0}
    pll_classes      = {k: [] for k in PLL_NAMES}
    flags            = []
    vendor_only      = []
    total_internal   = 0
    total_oe         = 0
    internal_courses = 0
    oe_courses       = 0
    country_counts   = {}

    for (course, loc_cat), data in classes.items():
        students = data["students"]
        capacity = data["capacity"]
        enrolled = students
        enroll_pct = round(enrolled / capacity * 100) if capacity else 0

        # Every Enrollment Database row is INTERNAL (bucket = source file, not
        # location). Distance-Learning / Santiago etc. are internal at "Other".
        location_counts[loc_cat] += students
        total_internal += students
        internal_courses += 1

        for country, count in data["countries"].items():
            country_counts[country] = country_counts.get(country, 0) + count

        pll_key, flagged = route_pll(course)

        if pll_key == "__vendor__":
            vendor_only.append(
                f"'{course}' ({loc_cat}, {students} students) "
                f"— vendor-led, counted in totals only"
            )
            continue

        if flagged:
            flags.append(
                f"UNMATCHED INTERNAL: '{course}' ({loc_cat}, {students} students) "
                f"— manual routing required"
            )
            continue

        pll_classes[pll_key].append({
            "name":        course,
            "capacity":    capacity,
            "enrolled":    enrolled,
            "enrollPct":   enroll_pct,
            "overEnrolled": enrolled > capacity if capacity else False,
            "locCat":      loc_cat,
            "students":    students,
            "is_oe":       False,
            "countries":   data["countries"],
        })

    return {
        "location_counts":  location_counts,
        "pll_classes":      pll_classes,
        "flags":            flags,
        "vendor_only":      vendor_only,
        "total_internal":   total_internal,
        "total_oe":         total_oe,
        "internal_courses": internal_courses,
        "oe_courses":       oe_courses,
        "country_counts":   country_counts,
    }


# ══════════════════════════════════════════════════════════════════════════
# CUSTOMER (OE / SS) PROCESSING
# ══════════════════════════════════════════════════════════════════════════

def process_customer_classes(path, bucket, week_start, week_end):
    """
    Read an OE (CMCustomerDemandList) or SS (ClassList) file and return the
    customer classes IN SESSION during the week. The two files share an
    identical 48-column layout; only the bucket label differs.

    - bucket: "OE" or "SS"
    - in-session: Delivery Start <= week_end AND Delivery End >= week_start
    - students: 'Contractual # of Students'
    - excludes Class Status == Cancelled
    - routed to a PLL by the Technology column (route_customer)

    Each returned class is tagged is_oe=True so build_pll_card lists it under
    the card's combined OE/SS block.
    """
    print(f"  Reading {bucket} file ({Path(path).name})…")
    headers, rows = load_sheet(path)

    col = {
        "title":    find_col(headers, "course title"),
        "classnm":  find_col(headers, "class name"),
        "tech":     find_col(headers, "technology"),
        "students": find_col(headers, "contractual # of students",
                             "contractual", "# of students"),
        "start":    find_col(headers, "delivery start date", "delivery start"),
        "finish":   find_col(headers, "delivery end date", "delivery end"),
        "location": find_col(headers, "location"),
        "city":     find_col(headers, "delivery city"),
        "country":  find_col(headers, "delivery country"),
        "status":   find_col(headers, "class status"),
    }

    def get(row, key):
        idx = col.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    pll_classes     = {k: [] for k in PLL_NAMES}
    flags           = []
    location_counts = {"HLC": 0, "BLC": 0, "KLC": 0, "Other": 0}
    delivery        = {}   # country -> students (for slide4 customer map)
    total_students  = 0
    total_classes   = 0

    for row in rows:
        name = str(get(row, "classnm") or get(row, "title") or "").strip()
        if not name:
            continue

        start  = to_date(get(row, "start"))
        finish = to_date(get(row, "finish"))
        if not start or not finish:
            continue
        if start > week_end or finish < week_start:
            continue  # not in session this week

        if str(get(row, "status") or "").strip().lower() == "cancelled":
            continue

        try:
            students = int(get(row, "students") or 0)
        except (TypeError, ValueError):
            students = 0
        if students <= 0:
            continue

        tech     = str(get(row, "tech") or "").strip()
        loc_cat  = classify_location(get(row, "location") or get(row, "city"))
        country  = canonical_country(get(row, "country"))

        total_students     += students
        total_classes      += 1
        location_counts[loc_cat] += students
        delivery[country]   = delivery.get(country, 0) + students

        pll_key, flagged = route_customer(tech, name)
        if flagged:
            flags.append(
                f"UNMATCHED {bucket}: '{name}' (tech='{tech}', "
                f"{students} students) — manual routing required"
            )
            continue

        pll_classes[pll_key].append({
            "name":      name,
            "students":  students,
            "is_oe":     True,
            "bucket":    bucket,
            "countries": {country: students},
            # placeholders (unused for OE/SS, which render as oeCourses):
            "capacity": 0, "enrolled": students, "enrollPct": 0,
            "overEnrolled": False, "locCat": loc_cat,
        })

    return {
        "pll_classes":     pll_classes,
        "flags":           flags,
        "location_counts": location_counts,
        "delivery":        delivery,
        "total_students":  total_students,
        "total_classes":   total_classes,
    }


# ══════════════════════════════════════════════════════════════════════════
# BOWLER PROCESSING
# ══════════════════════════════════════════════════════════════════════════

MONTH_ABBR = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
              7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}


def bowler_target_month(today):
    """Most recent complete month number (build month minus 1; Jan wraps to Dec)."""
    return 12 if today.month == 1 else today.month - 1


def parse_bowler_value(raw):
    """
    Normalize a raw Bowler Act cell to a percentage number (0-100, rounded
    to the nearest int), or None if it isn't usable (NA, TBD, blank, …).
    Values arrive as mixed decimals (0.91) and already-percentage numbers or
    strings (84.6, '*58%', '90%%') — strip stray characters, then treat
    <= 1.5 as a decimal fraction (×100).
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.upper() in ("NA", "N/A", "TBD"):
        return None
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s or s in ("-", "."):
        return None
    try:
        v = float(s)
    except ValueError:
        return None
    if v <= 1.5:
        v *= 100
    return round(v)


def load_bowler_sheet(path):
    """
    Load the raw Bowler Chart sheet and return:
      month_cols: {month_number (1-12): column_index}   (1Q/2Q/3Q/4Q skipped)
      kpi_rows:   {kpi_name_lower: {"py": row, "plan": row, "act": row}}
    The sheet's true header row (JAN…DEC) isn't the first row — load_sheet's
    headers/rows split doesn't apply here, so this reads the workbook directly.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    raw_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx = next(
        (i for i, r in enumerate(raw_rows)
         if r and any(str(c).strip().upper() == "JAN" for c in r if c is not None)),
        None
    )
    if header_idx is None:
        return {}, {}

    header = raw_rows[header_idx]
    month_cols = {}
    for col_i, cell in enumerate(header):
        label = str(cell).strip().upper() if cell else ""
        for m, abbr in enumerate(
            ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
             "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"], start=1
        ):
            if label == abbr:
                month_cols[m] = col_i

    # Each KPI spans 3 rows (PY / Plan / Act); only the first carries the
    # KPI name (col 2) — carry it forward across the triplet.
    kpi_rows = {}
    current_name = None
    for i in range(header_idx + 1, len(raw_rows)):
        row = raw_rows[i]
        name_cell = row[2] if len(row) > 2 else None
        if name_cell and str(name_cell).strip():
            current_name = str(name_cell).strip()
        row_type = str(row[7]).strip().lower() if len(row) > 7 and row[7] else ""
        if current_name and row_type in ("py", "plan", "act"):
            kpi_rows.setdefault(current_name.lower(), {})[row_type] = row

    return month_cols, kpi_rows


def find_kpi_act_row(kpi_rows, *keywords):
    """Find a KPI's Act row by substring match on its (lowercased) name."""
    for name_lower, rowset in kpi_rows.items():
        if any(kw in name_lower for kw in keywords):
            return rowset.get("act")
    return None


def bowler_month_value(month_cols, act_row, target_month):
    """
    Return (value, month_used) from an Act row, starting at target_month and
    walking backward through populated months until a usable numeric value
    is found. (None, None) if nothing is populated back to January.
    """
    if act_row is None:
        return None, None
    for m in range(target_month, 0, -1):
        col = month_cols.get(m)
        if col is None or col >= len(act_row):
            continue
        val = parse_bowler_value(act_row[col])
        if val is not None:
            return val, m
    return None, None


def bowler_overall_reason(kpis, overall):
    """Build the Bowler-card reason string from whichever KPI drives the overall RAG."""
    if overall == "green":
        return "All Bowler KPIs on plan"
    for key in BOWLER_CONFIG:
        k = kpis[key]
        if k["monthRag"] == overall and k["monthValue"] is not None:
            short = BOWLER_SHORT_NAMES.get(key, key)
            return f"{short} {k['monthLabel']} {k['monthValue']:g}% vs {k['plan']:g}% plan"
    return f"Bowler KPIs {overall}"


def process_bowler(month_cols, kpi_rows, target_month):
    """
    Extract Bowler KPI data for the most recent complete month, walking
    backward per-KPI to the latest populated month when the target month is
    blank/NA. Returns (kpis dict, overall_rag string, overall_reason string).
    """
    print(f"  Reading Bowler Chart ({MONTH_ABBR[target_month]} target)…")
    target_label = MONTH_ABBR[target_month]

    kpis = {}
    for key, cfg in BOWLER_CONFIG.items():
        kpis[key] = {
            "label":      cfg["label"],
            "ytdValue":   None,
            "ytdRag":     "amber",
            "monthLabel": target_label,
            "monthValue": None,
            "monthRag":   "amber",
            "plan":       cfg["plan"],
            "reverse":    cfg["reverse"],
        }

    act_row_for = {
        "ptsi":     find_kpi_act_row(kpi_rows, "skill improvement"),
        "timecard": find_kpi_act_row(kpi_rows, "timecard"),
        "instUtil": find_kpi_act_row(kpi_rows, "fully loaded and qualified instructors"),
    }

    for key, act_row in act_row_for.items():
        value, used_month = bowler_month_value(month_cols, act_row, target_month)
        if value is not None:
            kpis[key]["monthValue"] = value
            kpis[key]["monthLabel"] = MONTH_ABBR[used_month]
            kpis[key]["monthRag"]   = bowler_rag(key, value)

        # YTD = mean of populated Act months Jan..target_month (not the
        # walked-back month — YTD always spans the full year-to-date window).
        if act_row is not None:
            ytd_vals = []
            for m in range(1, target_month + 1):
                col = month_cols.get(m)
                if col is None or col >= len(act_row):
                    continue
                v = parse_bowler_value(act_row[col])
                if v is not None:
                    ytd_vals.append(v)
            if ytd_vals:
                kpis[key]["ytdValue"] = round(sum(ytd_vals) / len(ytd_vals))
                kpis[key]["ytdRag"]   = bowler_rag(key, kpis[key]["ytdValue"])

    overall = worst_rag(*[k["monthRag"] for k in kpis.values()])
    reason  = bowler_overall_reason(kpis, overall)
    return kpis, overall, reason


def process_safety_kpis(month_cols, kpi_rows, target_month):
    """
    Extract liveStop / readAcross fresh from the Bowler Chart's Act rows —
    no longer preserved from the previous board. Same month-walk-back logic
    as process_bowler(); readAcross in particular tends to be NA in the
    current month before Ops closes it out (falls back to the latest
    populated month, per CLAUDE.md).
    """
    live_row = find_kpi_act_row(kpi_rows, "live save rule compliance")
    read_row = find_kpi_act_row(kpi_rows, "read across closing rate")

    live_value, _ = bowler_month_value(month_cols, live_row, target_month)
    read_value, _ = bowler_month_value(month_cols, read_row, target_month)

    return {
        "liveStop":   {"value": live_value, "rag": safety_kpi_rag(live_value)},
        "readAcross": {"value": read_value, "rag": safety_kpi_rag(read_value)},
    }


# ══════════════════════════════════════════════════════════════════════════
# ACTION PLANS PROCESSING
# ══════════════════════════════════════════════════════════════════════════

def process_action_plans(headers, rows):
    """
    Extract action plan status counts from TOP-LEVEL PARENT rows only
    (live Smartsheet pull).

    The tracker is hierarchical (parent + child rows). Child rows are sub-tasks
    and must be excluded from the rollup — only top-level rows (Parent_Level 0)
    are reportable action plans.

    Status comes from TWO columns, both required:
      • 'Overall Status'  → Not Started / In Progress / On Hold / Complete / Cancelled
      • 'Current Status'  → Not Started / On Track / At Risk / Delayed

    Plus the 'At Risk' checkbox column and 'Current Finish' (drives overdue
    detection and the slide sort order).

    RAG:
      • red    — 2+ active APs Delayed or overdue
      • amber  — any active AP Delayed, overdue, or At Risk
      • green  — everything on track
    """
    print("  Processing Action Plan Tracker (parent rows only)…")

    col_plevel   = find_col(headers, "parent_level", "parent level")
    col_parentid = find_col(headers, "parentid", "parent id")
    col_ischild  = find_col(headers, "is child")
    col_overall  = find_col(headers, "overall status")
    col_current  = find_col(headers, "current status")
    col_title    = find_col(headers, "improvement", "description", "title")
    col_atrisk   = find_col(headers, "at risk")
    col_finish   = find_col(headers, "current finish")

    def get(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    def is_top_level(row):
        # Top-level action plan = Parent_Level 0 (preferred), else ParentID == 'Top',
        # else not flagged as a child. NOTE: the 'Is Parent' column means "has
        # children" (only ~115 rows), NOT "is top-level" — so it is deliberately
        # not used here. Values arrive as floats (0.0/1.0) via openpyxl.
        lvl = get(row, col_plevel)
        if lvl is not None and str(lvl).strip() != "":
            try:
                return float(lvl) == 0
            except (TypeError, ValueError):
                pass
        pid = get(row, col_parentid)
        if pid is not None and str(pid).strip() != "":
            return str(pid).strip().lower() == "top"
        child = get(row, col_ischild)
        if child is not None and str(child).strip() != "":
            try:
                return float(child) == 0
            except (TypeError, ValueError):
                return str(child).strip().lower() not in ("1", "true", "yes", "y")
        return True

    DELAYED_VOCAB   = {"delayed", "at risk"}
    EXCLUDE_OVERALL = {"complete", "cancelled", "on hold", ""}

    counts       = {"inProgress": 0, "delayed": 0, "notStarted": 0}
    items        = []
    overall_seen = {}
    current_seen = {}
    today        = date.today()
    n_delayed_or_overdue = 0      # drives red
    n_at_risk            = 0      # drives amber

    for row in rows:
        if not is_top_level(row):
            continue
        overall = str(get(row, col_overall) or "").strip()
        current = str(get(row, col_current) or "").strip()
        overall_seen[overall or "(blank)"] = overall_seen.get(overall or "(blank)", 0) + 1
        current_seen[current or "(blank)"] = current_seen.get(current or "(blank)", 0) + 1

        ol, cl = overall.lower(), current.lower()
        if ol in EXCLUDE_OVERALL:
            continue  # not an active reportable item

        finish  = to_date(get(row, col_finish))
        overdue = finish is not None and finish < today
        at_risk = bool(get(row, col_atrisk)) or cl == "at risk"

        if cl == "delayed" or overdue:
            n_delayed_or_overdue += 1
        if at_risk:
            n_at_risk += 1

        if cl in DELAYED_VOCAB:
            counts["delayed"] += 1
            bucket = "delayed"
        elif ol == "in progress":
            counts["inProgress"] += 1
            bucket = "inProgress"
        elif ol == "not started":
            counts["notStarted"] += 1
            bucket = "notStarted"
        else:
            continue

        if bucket in ("inProgress", "delayed"):
            title = str(get(row, col_title) or "").strip()
            if title:
                items.append({
                    "title":         title,
                    "currentStatus": current or overall,
                    "_finish":       finish,
                })

    # Slide content: active APs sorted by Current Finish ascending
    # (blank finish dates sort last).
    items.sort(key=lambda i: (i["_finish"] is None, i["_finish"] or today))
    for i in items:
        del i["_finish"]

    total = sum(counts.values())
    if n_delayed_or_overdue >= 2:
        ap_rag = "red"
    elif n_delayed_or_overdue or n_at_risk:
        ap_rag = "amber"
    else:
        ap_rag = "green"

    return {
        "deliveryTotal":    total,
        "inProgress":       counts["inProgress"],
        "delayed":          counts["delayed"],
        "notStarted":       counts["notStarted"],
        "atRisk":           n_at_risk,
        "inProgressItems":  items[:12],
        "rag":              ap_rag,
        "statusValues":     {"overall": overall_seen, "current": current_seen},
    }


# ══════════════════════════════════════════════════════════════════════════
# CAPEX PROCESSING
# ══════════════════════════════════════════════════════════════════════════

def process_capex(headers, rows):
    """
    Compute CapEx from the live Smartsheet (no longer carried forward).

    Row structure (hierarchy encoded in the 'Order' column = ancestors + 1):
      • Order 1  — year summary row ("CAPEX 2026 Overall") → budget totals
      • Order 2  — category subtotal rows                  → skipped
      • Order 3+ — actual line items                        → project list

    Filters: Year == 2026, Status != Cancelled.
    Budget total / spend come from the Order-1 summary row (Cost Estimate
    vs Spend = Actual + Projected). RAG stays amber per existing board logic.
    """
    print("  Processing CapEx (live Smartsheet)…")

    col_year     = find_col(headers, "year")
    col_order    = find_col(headers, "order")
    col_project  = find_col(headers, "project")
    col_cost     = find_col(headers, "cost estimate")
    col_priority = find_col(headers, "priority")
    col_category = find_col(headers, "category")
    col_status   = find_col(headers, "status")
    col_spend    = find_col(headers, "spend")

    def get(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    def fnum(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    CURRENT_YEAR = 2026.0
    DONE_STATUS  = "po received - all done"

    total_budget = 0.0
    total_spend  = 0.0
    items        = []
    by_category  = {}
    high_pri     = 0

    for row in rows:
        if fnum(get(row, col_year)) != CURRENT_YEAR:
            continue
        order = fnum(get(row, col_order))
        if order is None:
            continue

        if order == 1:                      # year summary row → budget totals
            total_budget += fnum(get(row, col_cost)) or 0
            total_spend  += fnum(get(row, col_spend)) or 0
            continue
        if order < 3:                       # category subtotal rows
            continue

        status = str(get(row, col_status) or "").strip()
        if status.lower() == "cancelled":
            continue

        project  = str(get(row, col_project) or "").strip()
        if not project:
            continue
        cost     = fnum(get(row, col_cost)) or 0
        priority = str(get(row, col_priority) or "").strip()
        # MULTIPICKLIST — attribute to the first listed category so category
        # bars never double-count a project.
        category = str(get(row, col_category) or "").split(",")[0].strip() or "Other"

        if priority == "High":
            high_pri += 1
        cat = by_category.setdefault(category, {"cost": 0, "count": 0})
        cat["cost"]  += cost
        cat["count"] += 1

        items.append({
            "project":  project,
            "cost":     cost,
            "category": category,
            "priority": priority,
            "_active":  bool(status) and status.lower() != DONE_STATUS,
        })

    if total_budget == 0:                   # no summary row → fall back to items
        total_budget = sum(i["cost"] for i in items)

    # Slide list: in-flight items (status set, not done), biggest cost first
    in_progress = sorted(
        (dict(i) for i in items if i["_active"]),
        key=lambda i: -i["cost"]
    )
    for i in in_progress:
        del i["_active"]

    return {
        "totalBudget":       round(total_budget),
        "totalSpend":        round(total_spend),
        "activeProjects":    len(items),
        "highPriorityCount": high_pri,
        "byCategory":        {k: {"cost": round(v["cost"]), "count": v["count"]}
                              for k, v in sorted(by_category.items(),
                                                 key=lambda kv: -kv[1]["cost"])},
        "inProgressItems":   in_progress[:12],
        "rag":               "amber",
    }


# ══════════════════════════════════════════════════════════════════════════
# XYLEME MODERNIZATION (Slide 1 card)
# ══════════════════════════════════════════════════════════════════════════

# Exam pipeline display buckets ← raw Exams Transfer Tracker statuses.
# Raw statuses not listed here (N/A, On Hold, blank) count toward the total
# only; the build summary prints them so drift is visible.
EXAM_STATUS_BUCKETS = {
    "not received":              "notReceived",
    "not started":               "notStarted",
    "in progress":               "inProgress",
    "sent for sme review":       "underReview",
    "sent for internal review":  "underReview",
    "internal review completed": "approved",
    "sme review completed":      "approved",
    "published":                 "published",
}


def process_xyleme(ss):
    """
    Build the slide-1 Xyleme card data from two Smartsheet trackers.

    • Modernization Tracker — depth-1 rows are modules ("Hierarchy 1.0");
      ring chart = Completed vs total, scroller = 10 most recently
      completed by End Date.
    • Exams Transfer Tracker — depth-1 rows are individual exams
      ("Level 1.0"); counted by status into pipeline buckets.
    """
    # ── Modernization Tracker ────────────────────────────────────────────
    headers, rows, depths = fetch_sheet_table(ss, "modernization",
                                              "Xyleme Modernization Tracker")
    col_project = find_col(headers, "project")
    col_status  = find_col(headers, "status")
    col_end     = find_col(headers, "end date")
    col_pl      = find_col(headers, "product line")

    def get(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    modules = [r for r, d in zip(rows, depths) if d == 1]
    completed = [r for r in modules
                 if str(get(r, col_status) or "").strip().lower() == "completed"]

    def end_key(r):
        d = to_date(get(r, col_end))
        return d.isoformat() if d else ""

    recently = sorted(completed, key=end_key, reverse=True)[:10]
    recently_published = []
    for r in recently:
        d = to_date(get(r, col_end))
        recently_published.append({
            "name": str(get(r, col_project) or "").strip(),
            "pl":   str(get(r, col_pl) or "").strip(),
            "date": d.isoformat() if d else "",
        })

    # ── Exams Transfer Tracker ───────────────────────────────────────────
    eh, erows, edepths = fetch_sheet_table(ss, "exams",
                                           "Xyleme Exams Transfer Tracker")
    col_estatus = find_col(eh, "status")

    exams = [r for r, d in zip(erows, edepths) if d == 1]
    counts = {"notReceived": 0, "notStarted": 0, "inProgress": 0,
              "underReview": 0, "approved": 0, "published": 0}
    unbucketed = {}
    for r in exams:
        raw = str(get(r, col_estatus) or "").strip()
        bucket = EXAM_STATUS_BUCKETS.get(raw.lower())
        if bucket:
            counts[bucket] += 1
        else:
            unbucketed[raw or "(blank)"] = unbucketed.get(raw or "(blank)", 0) + 1
    counts["total"] = len(exams)

    return {
        "modulesComplete":   len(completed),
        "modulesTotal":      len(modules),
        "exams":             counts,
        "recentlyPublished": recently_published,
        "_unbucketed":       unbucketed,   # build-summary diagnostics only
    }


# ══════════════════════════════════════════════════════════════════════════
# SAFETY
# ══════════════════════════════════════════════════════════════════════════

def calculate_safety_rag(weekly_incidents, bowler_overall, safety_kpis):
    """
    Safety RAG per CLAUDE.md: any KPI out of spec forces red \u2014 Life Saving
    Rules / Read Across compliance override the zero-incident green \u2014 then
    falls back to incident-count severity.
    """
    live = safety_kpis["liveStop"]
    read = safety_kpis["readAcross"]

    if live["rag"] == "red":
        return "red", f"Life Saving Rules compliance {live['value']}% vs 90% plan"
    if read["rag"] == "red":
        return "red", f"Read Across Closing Rate {read['value']}% vs 90% plan"
    if bowler_overall == "red":
        return "red", "KPI out of spec"
    if weekly_incidents == 0:
        return "green", "0 incidents this week \u2014 All safety KPIs on plan"
    if weekly_incidents <= 2:
        return "amber", f"{weekly_incidents} minor incident(s) \u2014 all KPIs in spec"
    return "red", f"{weekly_incidents} incidents this week"


# ══════════════════════════════════════════════════════════════════════════
# JSON ASSEMBLY
# ══════════════════════════════════════════════════════════════════════════

def build_pll_card(pll_key, class_list):
    """Assemble a slide2 PLL card object from processed class data."""
    courses    = []
    oe_courses = []
    total_students   = 0
    internal_students = 0
    oe_students      = 0

    for c in class_list:
        total_students += c["students"]
        if c["is_oe"]:
            oe_students += c["students"]
            top_country = (
                max(c["countries"], key=c["countries"].get)
                if c["countries"] else "Unknown"
            )
            oe_courses.append({
                "title":    c["name"],
                "students": c["students"],
                "country":  top_country,
            })
        else:
            internal_students += c["students"]
            courses.append({
                "name":        c["name"],
                "capacity":    c["capacity"],
                "enrolled":    c["enrolled"],
                "enrollPct":   c["enrollPct"],
                "overEnrolled":c["overEnrolled"],
                "locCat":      c["locCat"],
            })

    return {
        "name":             PLL_NAMES[pll_key],
        "totalClasses":     len(class_list),
        "totalStudents":    total_students,
        "internalStudents": internal_students,
        "oeStudents":       oe_students,
        "oeClasses":        len(oe_courses),
        "courses":          courses,
        "oeCourses":        oe_courses,
        "lookAhead30": {
            "label":        "Next 30 Days",
            "totalClasses": 0,
            "totalStudents":0,
            "intClasses":   0,
            "oeClasses":    0,
        },
    }


def build_pins(country_counts):
    """Build slide3 country pin array, sorted by student count descending."""
    pins = []
    for country, count in sorted(country_counts.items(), key=lambda x: -x[1]):
        lat, lon = country_coords(country)
        pins.append({
            "country":  country,
            "students": count,
            "lat":      lat,
            "lon":      lon,
        })
    return pins


def build_customer_delivery(oe_delivery, ss_delivery):
    """
    Build the slide4 customer-training-delivery markers (OE + SS) grouped by
    delivery country, sorted by student count descending. OE classes deliver at
    the learning centers; SS at customer sites — both are summed per country.
    """
    combined = {}
    for src in (oe_delivery, ss_delivery):
        for country, n in src.items():
            combined[country] = combined.get(country, 0) + n
    markers = []
    for country, n in sorted(combined.items(), key=lambda x: -x[1]):
        if n <= 0:
            continue
        lat, lon = country_coords(country)
        markers.append({"country": country, "lat": lat, "lon": lon, "count": n})
    return markers[:12]


# ══════════════════════════════════════════════════════════════════════════
# VALIDATION
# ══════════════════════════════════════════════════════════════════════════

def validate(board):
    """Run cross-checks and return list of validation errors."""
    errors = []
    s1 = board["slide1"]
    loc = s1["locationBreakdown"]

    loc_sum = sum(loc.values())
    if loc_sum != s1["weeklyTotal"]:
        errors.append(
            f"locationBreakdown sum {loc_sum} ≠ weeklyTotal {s1['weeklyTotal']}"
        )

    int_oe_sum = s1["internalStudents"] + s1["oeStudents"]
    if int_oe_sum != s1["weeklyTotal"]:
        errors.append(
            f"internalStudents {s1['internalStudents']} + oeStudents "
            f"{s1['oeStudents']} = {int_oe_sum} ≠ weeklyTotal {s1['weeklyTotal']}"
        )

    # slide3 is INTERNAL-only (country-of-origin map) — its total must match
    # internalStudents, not weeklyTotal.
    if board["slide3"]["totalStudents"] != s1["internalStudents"]:
        errors.append(
            f"slide3.totalStudents {board['slide3']['totalStudents']} "
            f"≠ internalStudents {s1['internalStudents']}"
        )

    return errors


# ══════════════════════════════════════════════════════════════════════════
# MAIN BUILD
# ══════════════════════════════════════════════════════════════════════════

def build(week_override=None):
    print("\n╔══════════════════════════════════════════╗")
    print("║  OFS Command Board — Weekly Build        ║")
    print("╚══════════════════════════════════════════╝\n")

    # ── Week ─────────────────────────────────────────────────────────────
    week_start, week_end = get_week(week_override)
    week_label = format_week_label(week_start, week_end)
    print(f"  Week: {week_label}\n")

    # ── File discovery (3 manual uploads) ────────────────────────────────
    print("Checking manual source files (data/)…")
    files, missing = discover_files()
    for key, path in files.items():
        print(f"  ✅  {key:<12}  {path.name}")
    for key in missing:
        print(f"  ❌  {key:<12}  NOT FOUND")

    if missing:
        print(f"\n⛔  Build stopped — {len(missing)} required file(s) missing.")
        print("    Upload the missing file(s) to data/ and run again.\n")
        sys.exit(1)

    print()

    # ── Load existing board (for preserving curated sections) ────────────
    existing = load_existing_board()

    # ── Pull live Smartsheet data ────────────────────────────────────────
    print("Pulling live Smartsheet data…")
    client = smartsheet_client()
    enr_headers, enr_rows, _ = fetch_sheet_table(client, "enrollment",
                                                 "Enrollment Database")
    ap_headers, ap_rows, _   = fetch_sheet_table(client, "actionplans",
                                                 "Action Plan Tracker")
    cx_headers, cx_rows, _   = fetch_sheet_table(client, "capex", "CapEx")
    xyleme = process_xyleme(client)
    print()

    # ── Process ──────────────────────────────────────────────────────────
    enr  = process_enrollment(enr_headers, enr_rows, week_start, week_end)   # Internal
    oe   = process_customer_classes(files["demand"],    "OE", week_start, week_end)  # Open Enrollment
    ss   = process_customer_classes(files["classlist"], "SS", week_start, week_end)  # Site Specific
    ap   = process_action_plans(ap_headers, ap_rows)
    cx   = process_capex(cx_headers, cx_rows)

    # KPIs / Bowler / Safety KPIs — derived fresh from the Bowler Chart every
    # build (no longer preserved from the previous board-data.json; the
    # month auto-selection + NA walk-back below replaced the unreliable
    # hardcoded-April parse that made preservation necessary).
    month_cols, kpi_rows = load_bowler_sheet(files["bowler"])
    target_month = bowler_target_month(date.today())
    kpis, bowler_overall, bowler_reason = process_bowler(month_cols, kpi_rows, target_month)
    safety_kpis = process_safety_kpis(month_cols, kpi_rows, target_month)

    # Safety — 0 new incidents assumed; Jim reviews weekly report manually
    weekly_incidents = 0
    safety_log = (existing.get("safetyLog") if existing and existing.get("safetyLog")
                  else list(SAFETY_LOG_BASE))   # carry forward cumulative log
    safety_rag, safety_reason = calculate_safety_rag(weekly_incidents, bowler_overall, safety_kpis)

    # ── Totals across the 3 buckets (Internal / OE / SS) ─────────────────
    internal_students = enr["total_internal"]
    oe_students       = oe["total_students"]
    ss_students       = ss["total_students"]
    customer_students = oe_students + ss_students          # OE + SS (oeStudents field)
    weekly_total      = internal_students + customer_students

    internal_courses  = enr["internal_courses"]
    customer_courses  = oe["total_classes"] + ss["total_classes"]

    # locationBreakdown = Internal (enrollment loc) + OE (at LC) + SS (at site)
    loc = {k: enr["location_counts"][k] + oe["location_counts"][k]
              + ss["location_counts"][k] for k in ("HLC", "BLC", "KLC", "Other")}

    internal_loc = enr["location_counts"]   # slide3 learning-center markers (internal only)

    all_flags = enr["flags"] + oe["flags"] + ss["flags"]

    # ── Build PLL cards (merge the 3 buckets per PLL) ─────────────────────
    merged = {k: enr["pll_classes"][k] + oe["pll_classes"][k] + ss["pll_classes"][k]
              for k in PLL_NAMES}
    plls = [build_pll_card(k, merged[k]) for k in PLL_ORDER]

    # Preserve each PLL's 30-day look-ahead from the existing board (not yet
    # computed from source — carry forward by PLL name).
    if existing and "slide2" in existing:
        prev_cards = {p.get("name"): p for p in existing["slide2"].get("plls", [])
                      if isinstance(p, dict) and "name" in p}
        for card in plls:
            prev = prev_cards.get(card["name"])
            if prev and "lookAhead30" in prev:
                card["lookAhead30"] = prev["lookAhead30"]

    plls.append({"name": "__WAG__"})   # sentinel — must be last

    # ── Country pins ─────────────────────────────────────────────────────
    pins = build_pins(enr["country_counts"])

    # ── Assemble JSON ────────────────────────────────────────────────────
    board = {
        "meta": {
            "specVersion": "v2.2",
            "buildDate":   date.today().isoformat(),
            "weekOf":      week_label,
        },
        "slide1": {
            "weekOf":           week_label,
            "weeklyTotal":      weekly_total,
            "internalStudents": internal_students,
            "oeStudents":       customer_students,
            "internalCourses":  internal_courses,
            "oeCourses":        customer_courses,
            "locationBreakdown": loc,
            "safetyRag":        safety_rag,
            "minorIncidents":   weekly_incidents,
            "ragReason":        safety_reason,
            "safetyKPIs": safety_kpis,
            "kpis": kpis,
            "overallRAGs": {
                "bowler": {
                    "rag":    bowler_overall,
                    "label":  "Bowler KPIs",
                    "reason": bowler_reason,
                },
                "safety": {
                    "rag":    safety_rag,
                    "label":  "Safety",
                    "reason": safety_reason,
                },
                "actionPlans": {
                    "rag":    ap["rag"],
                    "label":  "Action Plans",
                    "reason": f"{ap['delayed']} delayed / {ap['inProgress']} in progress",
                },
                "capex": {
                    "rag":    cx["rag"],
                    "label":  "CapEx",
                    "reason": f"{cx['activeProjects']} active projects",
                },
            },
        },
        "safetyLog": safety_log,
        "slide2": {"plls": plls},
        "slide3": {
            # Internal only \u2014 student country-of-origin + learning-center markers
            "totalCountries": len(enr["country_counts"]),
            "totalStudents":  internal_students,
            "learningCenters": [
                {"name": "HLC \u2014 Houston", "lat": 29.7604, "lon": -95.3698, "students": internal_loc["HLC"]},
                {"name": "BLC \u2014 Birr",    "lat": 47.36,   "lon":   8.04,   "students": internal_loc["BLC"]},
                {"name": "KLC \u2014 Kuwait",  "lat": 29.3117, "lon":  47.4818, "students": internal_loc["KLC"]},
            ],
            "pins": pins,
        },
        "slide4": {
            # Customer training (OE + SS) by delivery location
            "weekStudents": customer_students,
            "weekClasses":  customer_courses,
            "weekDelivery": build_customer_delivery(oe["delivery"], ss["delivery"]),
        },
        "actionPlans": {
            "deliveryTotal":   ap["deliveryTotal"],
            "inProgress":      ap["inProgress"],
            "delayed":         ap["delayed"],
            "notStarted":      ap["notStarted"],
            "inProgressItems": ap["inProgressItems"],
        },
        "capex": {
            "totalBudget":        cx["totalBudget"],
            "totalSpend":         cx.get("totalSpend", 0),
            "activeProjects":     cx["activeProjects"],
            "highPriorityCount":  cx["highPriorityCount"],
            "byCategory":         cx["byCategory"],
            "inProgressItems":    cx["inProgressItems"],
        },
        "xyleme": {
            "modulesComplete":   xyleme["modulesComplete"],
            "modulesTotal":      xyleme["modulesTotal"],
            "exams":             xyleme["exams"],
            "recentlyPublished": xyleme["recentlyPublished"],
        },
    }

    # ── Validate ─────────────────────────────────────────────────────────
    errors = validate(board)
    if errors:
        print("⛔  Validation errors — board-data.json NOT written:")
        for e in errors:
            print(f"    → {e}")
        sys.exit(1)

    # ── Write ────────────────────────────────────────────────────────────
    out = Path("board-data.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(board, f, indent=2, ensure_ascii=False)

    # ── Build summary ────────────────────────────────────────────────────
    bar = "═" * 52
    print(f"\n{bar}")
    print("  BUILD SUMMARY — review before committing")
    print(bar)
    print(f"  Week:          {week_label}")
    print(f"  Total students:{weekly_total:>5}   "
          f"(Internal {internal_students} / OE {oe_students} / SS {ss_students})")
    print(f"  Total classes: {internal_courses + customer_courses:>5}   "
          f"(Internal {internal_courses} / OE {oe['total_classes']} / SS {ss['total_classes']})")
    print(f"  HLC / BLC / KLC / Other:  "
          f"{loc['HLC']} / {loc['BLC']} / {loc['KLC']} / {loc['Other']}")
    print()
    print("  PLL Summary  (Internal + OE + SS):")
    for key in PLL_ORDER:
        classes = merged[key]
        n_cls = len(classes)
        n_stu = sum(c["students"] for c in classes)
        n_int = sum(c["students"] for c in classes if not c["is_oe"])
        n_cust = n_stu - n_int
        print(f"    {PLL_NAMES[key]:<24} {n_cls:>2} class(es)  {n_stu:>3} students"
              f"   (INT {n_int} / OE+SS {n_cust})")
    print()
    print(f"  Safety RAG:    {safety_rag.upper():<8} — {safety_reason}")
    print(f"  Bowler RAG:    {bowler_overall.upper()}")
    print(f"  Action Plans:  {ap['rag'].upper():<8} — {ap['deliveryTotal']} active "
          f"({ap['inProgress']} in progress / {ap['delayed']} delayed / "
          f"{ap['notStarted']} not started / {ap.get('atRisk', 0)} at risk)  "
          f"[parent rows only]")
    print(f"  CapEx:         {cx['activeProjects']} projects / "
          f"${cx['totalBudget']:,.0f} budget / ${cx.get('totalSpend', 0):,.0f} spend "
          f"[live Smartsheet]")
    print(f"  Xyleme:        {xyleme['modulesComplete']}/{xyleme['modulesTotal']} "
          f"modules complete / {xyleme['exams']['total']} exams in pipeline")
    if xyleme.get("_unbucketed"):
        ub = ", ".join(f"{k}={v}" for k, v in sorted(xyleme["_unbucketed"].items()))
        print(f"      ℹ️  exam statuses outside pipeline buckets (total-only): {ub}")
    print()

    # ── Preserved / verify-manually flags ────────────────────────────────
    print("  ⚠️   VERIFY MANUALLY — sections carried forward, not derived from source:")
    print("      → PLL lookAhead30: preserved from previous board (not yet computed)")
    print()

    # ── Action-Plan status diagnostics (confirm the parent filter) ────────
    sv = ap.get("statusValues", {})
    if sv:
        print("  📋  Action-Plan status values among PARENT rows "
              "(confirm filter is right):")
        ov = ", ".join(f"{k}={v}" for k, v in sorted(sv.get("overall", {}).items()))
        cu = ", ".join(f"{k}={v}" for k, v in sorted(sv.get("current", {}).items()))
        print(f"      Overall Status:  {ov}")
        print(f"      Current Status:  {cu}")
        print(f"      → counted: {ap['inProgress']} inProg + {ap['delayed']} delayed "
              f"+ {ap['notStarted']} notStarted = {ap['deliveryTotal']} "
              f"(prior board: 11 / 32 / 48 = 91)")
        print()

    if enr.get("vendor_only"):
        print(f"  ℹ️   VENDOR-LED — {len(enr['vendor_only'])} class(es) counted "
              "in totals only (no PLL card):")
        for note in enr["vendor_only"]:
            print(f"      → {note}")
        print()

    if all_flags:
        print(f"  ⚠️   FLAGGED — {len(all_flags)} unmatched class(es) need manual routing:")
        for flag in all_flags:
            print(f"      → {flag}")
        print()

    print(f"  board-data.json written  ({out.stat().st_size:,} bytes)")
    print()
    print("  ✅  Review complete. Tell Claude Code 'approved' or 'commit it'")
    print(f"      to run: git commit -m \"Week of {week_label} — "
          f"{weekly_total} students, "
          f"{internal_courses + customer_courses} classes\"")
    print(f"{bar}\n")

    return board, all_flags


# ══════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="OFS Command Board — weekly data builder"
    )
    parser.add_argument(
        "--week", metavar="YYYY-MM-DD",
        help="Monday of the target week (default: WEEK_OF env var, "
             "else current week)"
    )
    args = parser.parse_args()
    week = args.week or os.environ.get("WEEK_OF", "").strip() or None
    build(week_override=week)
